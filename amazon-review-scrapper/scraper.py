#!/usr/bin/env python3
"""
Amazon India Review Scraper — Version 1
========================================
Scrapes all written reviews from amazon.in for a given ASIN.
Uses a persistent Chrome session — sign in once with your throwaway
amazon.in account; the session is reused for every future run.

Usage:
    # First run — Chrome opens, sign in manually
    python3 scraper.py B0CVXQ9SKB

    # Filter by star rating
    python3 scraper.py B0CVXQ9SKB --stars 5

    # Run headless (only after first sign-in)
    python3 scraper.py B0CVXQ9SKB --headless
"""

import argparse
import os
import sys
import time
import random

from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth
from bs4 import BeautifulSoup
import pandas as pd

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(SCRIPT_DIR, "exports")
PROFILE_DIR = os.path.join(SCRIPT_DIR, ".browser_profile")

# ── Constants ─────────────────────────────────────────────────────────────────
TLD               = "in"                   # amazon.in — fixed for V1
LOGIN_WAIT_SECS   = 300                    # 5 minutes to sign in
STAR_ROTATION     = ["five_star", "four_star", "three_star", "two_star", "one_star"]
STAR_CHOICES      = {
    "all": STAR_ROTATION,
    "5":   ["five_star"],
    "4":   ["four_star"],
    "3":   ["three_star"],
    "2":   ["two_star"],
    "1":   ["one_star"],
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def build_url(asin, star_filter):
    return (
        f"https://www.amazon.{TLD}/product-reviews/{asin}/"
        f"?ie=UTF8&reviewerType=all_reviews&filterByStar={star_filter}&pageNumber=1"
    )


def safe_content(page):
    """Return page HTML, waiting for any in-flight navigation to settle."""
    for _ in range(3):
        try:
            page.wait_for_load_state("domcontentloaded", timeout=10000)
            return page.content()
        except Exception:
            time.sleep(1)
    return page.content()


def is_blocked(html):
    soup  = BeautifulSoup(html, "lxml")
    title = soup.title.get_text(strip=True).lower() if soup.title else ""
    has_captcha    = bool(soup.select_one("form[action*='validateCaptcha']"))
    is_signin      = "sign-in" in title or "sign in" in title
    is_verify      = "ax/claim" in title or "authentication" in title
    is_passkey     = "passkey" in title
    return has_captcha or is_signin or is_verify or is_passkey


def get_total_review_count(html):
    soup = BeautifulSoup(html, "lxml")
    el   = soup.select_one('[data-hook="total-review-count"]')
    if el:
        try:
            return int(el.get_text(strip=True).replace(",", "").split()[0])
        except (ValueError, IndexError):
            pass
    return None


def wait_for_login(page, target_url):
    """Pause until the user finishes signing in / solves the CAPTCHA."""
    print("\n  ⚠️  Amazon needs your attention in the Chrome window.")
    print("  → Could be: sign-in, CAPTCHA, or a Passkey setup prompt.")
    print("  → Complete it or dismiss it, then wait for reviews to load.")
    print(f"  → Waiting up to {LOGIN_WAIT_SECS // 60} minutes...\n")

    start = time.time()
    while time.time() - start < LOGIN_WAIT_SECS:
        try:
            html  = safe_content(page)
            if not is_blocked(html):
                soup  = BeautifulSoup(html, "lxml")
                title = (soup.title.get_text(strip=True) if soup.title else "").lower()
                if soup.select('[data-hook="review"]') or "customer reviews" in title:
                    print("  ✓ Signed in. Continuing scrape...\n")
                    return True
        except Exception:
            pass
        time.sleep(2)

    print("  ✗ Timed out waiting for sign-in.")
    return False


# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_reviews(html):
    soup    = BeautifulSoup(html, "lxml")
    reviews = []

    for div in soup.select('[data-hook="review"]'):
        r = {}
        r["_id"] = div.get("id") or ""

        el = div.select_one("span.a-profile-name")
        r["Reviewer"] = el.get_text(strip=True) if el else ""

        el = div.select_one('i[data-hook="review-star-rating"], i[data-hook="cmps-review-star-rating"]')
        if el:
            try:
                r["Rating"] = float(el.get_text(strip=True).split()[0])
            except (ValueError, IndexError):
                r["Rating"] = ""
        else:
            r["Rating"] = ""

        el = div.select_one('a[data-hook="review-title"] span, span[data-hook="review-title"]')
        if el:
            parent = el.parent
            if parent and parent.name == "a":
                spans     = parent.select("span")
                r["Title"] = spans[-1].get_text(strip=True) if spans else el.get_text(strip=True)
            else:
                r["Title"] = el.get_text(strip=True)
        else:
            r["Title"] = ""

        el = div.select_one('span[data-hook="review-date"]')
        r["Date"] = el.get_text(strip=True) if el else ""

        el = div.select_one('span[data-hook="review-body"]')
        r["Review Text"] = el.get_text(strip=True) if el else ""

        el = div.select_one('span[data-hook="avp-badge"]')
        r["Verified Purchase"] = "Yes" if el else "No"

        el = div.select_one('span[data-hook="helpful-vote-statement"]')
        r["Helpful Votes"] = el.get_text(strip=True) if el else "0"

        reviews.append(r)

    return reviews


# ── Core scrape logic ─────────────────────────────────────────────────────────

def scrape_filter(page, asin, star_filter, max_batches, seen_ids, collected):
    """
    Navigate to the star-filtered review page, then repeatedly click
    'Show more reviews' (Amazon India's AJAX loader) until exhausted.
    """
    print(f"\n─── {star_filter.replace('_', ' ').title()} ───")

    url = build_url(asin, star_filter)

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=25000)
    except Exception as e:
        print(f"  ✗ Navigation failed: {e}")
        return

    try:
        page.wait_for_selector('[data-hook="review"]', timeout=8000)
    except Exception:
        pass  # may be sign-in page — handled below

    html = safe_content(page)

    if is_blocked(html):
        if not wait_for_login(page, url):
            return
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=25000)
            page.wait_for_selector('[data-hook="review"]', timeout=8000)
        except Exception:
            pass
        html = safe_content(page)
        if is_blocked(html):
            print("  ✗ Still blocked. Skipping this filter.")
            return

    total = get_total_review_count(html)
    if total:
        print(f"  Product has {total} total ratings.")

    for batch in range(1, max_batches + 1):
        html = safe_content(page)

        # Mid-session block check (e.g. session expired mid-scrape).
        if is_blocked(html):
            if not wait_for_login(page, url):
                return
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=25000)
                page.wait_for_selector('[data-hook="review"]', timeout=8000)
            except Exception:
                pass
            html = safe_content(page)

        page_reviews = parse_reviews(html)

        if batch == 1 and not page_reviews:
            soup  = BeautifulSoup(html, "lxml")
            title = soup.title.get_text(strip=True) if soup.title else "unknown"
            print(f"  ✗ No reviews found (page title: '{title}')")
            return

        new_reviews = []
        for r in page_reviews:
            key = r["_id"] if r["_id"] else f"{r['Reviewer']}|{r['Date']}|{r['Review Text'][:80]}"
            if key not in seen_ids:
                seen_ids.add(key)
                new_reviews.append(r)
                collected.append(r)

        if new_reviews:
            print(f"  Batch {batch}/{max_batches} — +{len(new_reviews)} new  (total so far: {len(collected)})")

        # Click 'Show more reviews' — Amazon India's AJAX pagination.
        show_more = page.query_selector('[data-hook="show-more-button"]')
        if not show_more:
            print("  → No more reviews. Filter exhausted.")
            break

        prev_count = len(page.query_selector_all('[data-hook="review"]'))
        show_more.click()
        try:
            page.wait_for_function(
                f'document.querySelectorAll(\'[data-hook="review"]\').length > {prev_count}',
                timeout=10000,
            )
        except Exception:
            pass
        time.sleep(random.uniform(1, 2))


def scrape(asin, star_filters, max_batches, headless):
    collected = []
    seen_ids  = set()
    stealth   = Stealth()

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            headless=headless,
            channel="chrome",
            viewport={"width": 1280, "height": 900},
            locale="en-US",
            args=["--disable-blink-features=AutomationControlled"],
        )
        stealth.apply_stealth_sync(context)
        page = context.pages[0] if context.pages else context.new_page()

        for star_filter in star_filters:
            scrape_filter(page, asin, star_filter, max_batches, seen_ids, collected)

        context.close()

    for r in collected:
        r.pop("_id", None)
    return collected


# ── Output ────────────────────────────────────────────────────────────────────

def save_excel(reviews, asin):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Growisto brand colours (ARGB) ─────────────────────────────────────────
    TEAL_BLUE   = "FF367588"   # headers, title bar
    POWDER_BLUE = "FFB8DBD9"   # borders, summary row
    RAISIN      = "FF1D1D20"   # body text
    CULTURED    = "FFF6F6F4"   # data row backgrounds
    WHITE       = "FFFFFFFF"   # header text
    FONT        = "Poppins"

    # ── Style factories ───────────────────────────────────────────────────────
    thin_border = Border(
        left   = Side(style="thin", color=POWDER_BLUE[2:]),
        right  = Side(style="thin", color=POWDER_BLUE[2:]),
        top    = Side(style="thin", color=POWDER_BLUE[2:]),
        bottom = Side(style="thin", color=POWDER_BLUE[2:]),
    )

    def header_style(cell):
        cell.fill      = PatternFill("solid", fgColor=TEAL_BLUE)
        cell.font      = Font(name=FONT, color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = thin_border

    def data_style(cell, bold=False):
        cell.fill      = PatternFill("solid", fgColor=CULTURED)
        cell.font      = Font(name=FONT, color=RAISIN, bold=bold, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border    = thin_border

    # ── Build & write ─────────────────────────────────────────────────────────
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    df       = pd.DataFrame(reviews)
    filename = f"reviews_{asin}_in.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Reviews")
        ws = writer.sheets["Reviews"]

        # ── Title banner (row 1 — inserted above headers) ─────────────────────
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value     = f"Amazon India Reviews  ·  ASIN: {asin}  ·  Growisto Internal Tool"
        title_cell.fill      = PatternFill("solid", fgColor=TEAL_BLUE)
        title_cell.font      = Font(name=FONT, color=WHITE, bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Column headers (now row 2) ─────────────────────────────────────────
        ws.row_dimensions[2].height = 22
        for col_idx in range(1, len(df.columns) + 1):
            header_style(ws.cell(row=2, column=col_idx))

        # ── Data rows (row 3 onward) ───────────────────────────────────────────
        for row_idx in range(3, len(df) + 3):
            ws.row_dimensions[row_idx].height = 60
            for col_idx in range(1, len(df.columns) + 1):
                data_style(ws.cell(row=row_idx, column=col_idx))

        # ── Rating column: colour-code by star value ───────────────────────────
        try:
            rating_col = list(df.columns).index("Rating") + 1
            star_colours = {
                5.0: "FF2E7D32",   # dark green
                4.0: "FF558B2F",   # light green
                3.0: "FFF9A825",   # amber
                2.0: "FFEF6C00",   # orange
                1.0: "FFC62828",   # red
            }
            for row_idx, rating in enumerate(df["Rating"], start=3):
                cell = ws.cell(row=row_idx, column=rating_col)
                colour = star_colours.get(float(rating) if rating != "" else 0)
                if colour:
                    cell.fill = PatternFill("solid", fgColor=colour)
                    cell.font = Font(name=FONT, color=WHITE, bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="top")
        except (ValueError, TypeError):
            pass

        # ── Column widths ──────────────────────────────────────────────────────
        col_widths = {
            "Reviewer"         : 20,
            "Rating"           : 8,
            "Title"            : 30,
            "Date"             : 22,
            "Review Text"      : 60,
            "Verified Purchase": 14,
            "Helpful Votes"    : 14,
        }
        for col_idx, col_name in enumerate(df.columns, 1):
            width = col_widths.get(col_name, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Freeze panes under title + header ─────────────────────────────────
        ws.freeze_panes = "A3"

        # ── Auto-filter on header row ──────────────────────────────────────────
        ws.auto_filter.ref = (
            f"A2:{get_column_letter(len(df.columns))}2"
        )

    return filepath


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Amazon India Review Scraper — V1"
    )
    parser.add_argument("asin",
        help="10-character Amazon ASIN (e.g. B0CVXQ9SKB)")
    parser.add_argument("--pages", "-p", type=int, default=100,
        help="Max 'Show more' clicks per star filter (default: 100 ≈ 1000 reviews)")
    parser.add_argument("--stars", "-s", default="all",
        choices=list(STAR_CHOICES.keys()),
        help="Star filter: all (default) or 1–5")
    parser.add_argument("--headless", action="store_true",
        help="Hide Chrome window (only after first sign-in)")
    args = parser.parse_args()

    asin = args.asin.strip().upper()
    if len(asin) != 10:
        print(f"Error: ASIN must be 10 characters — got '{asin}' ({len(asin)} chars)")
        sys.exit(1)

    star_filters   = STAR_CHOICES[args.stars]
    profile_exists = os.path.isdir(PROFILE_DIR)

    print(f"\nAmazon India Review Scraper — V1")
    print(f"  ASIN    : {asin}")
    print(f"  Stars   : {args.stars}")
    print(f"  Max batches per filter: {args.pages}")
    print(f"  Headless: {args.headless}")
    print(f"  Session : {'saved ✓' if profile_exists else 'none — Chrome will open for sign-in'}")

    if args.headless and not profile_exists:
        print("\n  ⚠️  No saved session — remove --headless so you can sign in first.")
        sys.exit(1)

    print()

    reviews = scrape(asin, star_filters, args.pages, headless=args.headless)

    if not reviews:
        print("\nNo reviews collected.")
        sys.exit(1)

    filepath = save_excel(reviews, asin)
    print(f"\n✓ Done — {len(reviews)} reviews saved to:\n  {filepath}")


if __name__ == "__main__":
    main()
